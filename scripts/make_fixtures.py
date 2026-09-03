"""
Generates realistic .xlsx test fixtures for the Stock Scanner app using
openpyxl (an independent, real-world Excel library) — NOT the app's own
writer — so the app's reader is validated against genuine Excel output,
not just against itself.
"""
import openpyxl
from openpyxl.styles import Font
import os

OUT_DIR = r"..\app\src\test\resources\fixtures"
os.makedirs(OUT_DIR, exist_ok=True)

# --- Fixture 1: the "normal" file, columns in a DIFFERENT order than the
# app's writer uses, plus an extra irrelevant column, blank locations,
# a numeric-looking sku kept as TEXT (leading zeros must survive), a merged
# header-ish styling, and a completely blank trailing row (common Excel export
# artifact) that must be skipped.
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"
# Deliberately not in the app's canonical A,B,C,D order, and with an extra
# "הערות" column the app must ignore.
headers = ["ברקוד", "מקט", "הערות", "תאור", "מיקום"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True)

rows = [
    ("7290012345678", "ABC-123", "", "פילטר שמן טויוטה", ""),
    ("7290000000002", "0001234", "מלאי נמוך", "מסנן אוויר", "A-01-01"),
    ("729123456789", "PRD_00123", "", "  מצבר   12V  ", "A-01-02"),  # extra internal whitespace
    ("", "A12-B45", "", "פנס אחורי ימין", ""),  # blank barcode
]
for r_idx, (barcode, sku, note, desc, loc) in enumerate(rows, start=2):
    ws.cell(row=r_idx, column=1, value=barcode)
    ws.cell(row=r_idx, column=2, value=sku)
    ws.cell(row=r_idx, column=3, value=note)
    ws.cell(row=r_idx, column=4, value=desc)
    ws.cell(row=r_idx, column=5, value=loc)

# A fully blank trailing row (openpyxl/Excel sometimes leaves these behind)
for col in range(1, 6):
    ws.cell(row=len(rows) + 2, column=col, value=None)

wb.save(os.path.join(OUT_DIR, "sample_normal.xlsx"))

# --- Fixture 2: duplicate skus and duplicate barcodes, to validate the
# de-duplication + warning-count logic.
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Sheet1"
headers2 = ["מקט", "תאור", "ברקוד", "מיקום"]
for col, h in enumerate(headers2, start=1):
    ws2.cell(row=1, column=col, value=h)

rows2 = [
    ("SKU-1", "מוצר ראשון (ישן)", "1111", ""),
    ("SKU-2", "מוצר שני", "2222", ""),
    ("SKU-1", "מוצר ראשון (מעודכן)", "1111", ""),  # duplicate sku -> last wins
    ("SKU-3", "מוצר שלישי", "2222", ""),  # duplicate barcode with SKU-2
]
for r_idx, (sku, desc, barcode, loc) in enumerate(rows2, start=2):
    ws2.cell(row=r_idx, column=1, value=sku)
    ws2.cell(row=r_idx, column=2, value=desc)
    ws2.cell(row=r_idx, column=3, value=barcode)
    ws2.cell(row=r_idx, column=4, value=loc)

wb2.save(os.path.join(OUT_DIR, "sample_duplicates.xlsx"))

# --- Fixture 3: missing a required column, to validate the clear error message.
wb3 = openpyxl.Workbook()
ws3 = wb3.active
headers3 = ["מקט", "תאור", "מיקום"]  # ברקוד missing
for col, h in enumerate(headers3, start=1):
    ws3.cell(row=1, column=col, value=h)
ws3.cell(row=2, column=1, value="X1")
ws3.cell(row=2, column=2, value="מוצר")
ws3.cell(row=2, column=3, value="")
wb3.save(os.path.join(OUT_DIR, "sample_missing_column.xlsx"))

# --- Fixture 4: multi-location columns ("מיקום", "מיקום 2", "מיקום 3", ...),
# in a non-contiguous header order, to validate that the reader combines them
# regardless of physical column position.
wb4 = openpyxl.Workbook()
ws4 = wb4.active
# Note the numbered location columns are deliberately not adjacent to each
# other or to "מיקום", and "מיקום 3" appears before "מיקום 2".
headers4 = ["מקט", "מיקום 3", "תאור", "ברקוד", "מיקום", "מיקום 2"]
for col, h in enumerate(headers4, start=1):
    ws4.cell(row=1, column=col, value=h)

rows4 = [
    # sku, מיקום 3, תאור, ברקוד, מיקום, מיקום 2
    ("MULTI-1", "C-03-01", "מוצר בשלושה מקומות", "111", "A-01-05", "B-02-01"),
    ("SINGLE-1", "", "מוצר במקום אחד", "222", "A-01-06", ""),
    ("NONE-1", "", "מוצר בלי מיקום", "333", "", ""),
]
for r_idx, (sku, loc3, desc, barcode, loc1, loc2) in enumerate(rows4, start=2):
    ws4.cell(row=r_idx, column=1, value=sku)
    ws4.cell(row=r_idx, column=2, value=loc3)
    ws4.cell(row=r_idx, column=3, value=desc)
    ws4.cell(row=r_idx, column=4, value=barcode)
    ws4.cell(row=r_idx, column=5, value=loc1)
    ws4.cell(row=r_idx, column=6, value=loc2)

wb4.save(os.path.join(OUT_DIR, "sample_multi_location.xlsx"))

print("Fixtures written to", OUT_DIR)
print(os.listdir(OUT_DIR))
